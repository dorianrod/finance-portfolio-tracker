#!/usr/bin/env python3
"""Build or update the dated allocation workbook from research sheets.

Reads the markdown reports produced by the allocation-update skill's
research workflow (one per asset, under
<data-dir>/input/allocations/{date}/*.md) and writes their "New values"
into <data-dir>/input/allocations/{date}.xlsx, starting from the most
recent existing workbook so untouched assets carry their values forward.

Usage:
    python build_allocation_xlsx.py [--date YYYY-MM-DD]
                                     [--data-dir PATH]
                                     [--allocations-dir PATH]
                                     [--sheets-dir PATH]
                                     [--sheet path/to/one-report.md]
                                     [--dry-run]

Output:
    JSON synthesis on stdout: which assets were created/updated, every
    changed cell's old/new value, skipped cells, and warnings.
"""

import argparse
import json
import os
import re
import shutil
import sys
import unicodedata
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(
        json.dumps(
            {
                "error": "openpyxl is not installed. "
                "Run: pip install pandas openpyxl"
            }
        )
    )
    sys.exit(1)

CATEGORY_RANGES_0INDEXED: dict[str, tuple[int, int]] = {
    "geo": (2, 12),
    "secteur": (12, 26),
    "currency": (26, 30),
    "classe": (30, 39),
}

HEADING_TO_CATEGORY = {
    "geographic": "geo",
    "sector": "secteur",
    "currency": "currency",
    "asset class": "classe",
}

NAME_COL_0 = 0  # nom_placement
REFERENCE_HEADER = "reference"
HEADER_ROW_1INDEXED = 2  # openpyxl row holding column headers
DATA_START_ROW_1INDEXED = 3  # first data row


def resolve_data_dir(cli_value: str | None) -> Path:
    if cli_value:
        return Path(cli_value).expanduser().resolve()
    env_value = os.environ.get("FINANCE_DATA_DIR")
    if env_value:
        return Path(env_value).expanduser().resolve()
    cwd = Path.cwd()
    if (cwd / "input").is_dir():
        return cwd
    return cwd / "data"


def _norm_key(s: str) -> str:
    s = " ".join(str(s).split()).lower()
    return "".join(
        c
        for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def _parse_table_row(line: str) -> list[str]:
    cells = line.strip().strip("|").split("|")
    return [c.strip() for c in cells]


def parse_report(text: str) -> dict:
    """Parse one Workflow A markdown report.

    Returns {"asset_name": str, "sources": [url, ...],
             "values": {category: {column: raw_value_str}}}.
    """
    first_line = next((ln for ln in text.splitlines() if ln.strip()), "")
    m = re.match(r"^Asset name:\s*(.+)$", first_line.strip())
    if not m:
        raise ValueError(
            "First non-empty line must be 'Asset name: <name>'"
        )
    asset_name = m.group(1).strip()

    sources_match = re.search(
        r"^# Sources\s*$(.*?)(?=^# |\Z)", text, re.MULTILINE | re.DOTALL
    )
    sources: list[str] = []
    if sources_match:
        for url_match in re.finditer(
            r"\[[^\]]*\]\(([^)]+)\)", sources_match.group(1)
        ):
            url = url_match.group(1).strip()
            if url and url not in sources:
                sources.append(url)

    exposures_match = re.search(
        r"^# Exposures\s*$(.*)", text, re.MULTILINE | re.DOTALL
    )
    values: dict[str, dict[str, str]] = {}
    if exposures_match:
        block = exposures_match.group(1)
        sections = re.split(r"(?m)^## ", block)
        for section in sections[1:]:
            heading_line, _, rest = section.partition("\n")
            heading = heading_line.strip().lower()
            category = HEADING_TO_CATEGORY.get(heading)
            if category is None:
                continue
            table_lines = [
                ln for ln in rest.splitlines() if ln.strip().startswith("|")
            ]
            if len(table_lines) < 3:
                continue
            header_cells = _parse_table_row(table_lines[0])
            new_values_row = next(
                (ln for ln in table_lines[2:] if "new values" in ln.lower()),
                None,
            )
            if new_values_row is None:
                continue
            value_cells = _parse_table_row(new_values_row)
            cat_values: dict[str, str] = {}
            for col_name, raw_val in zip(
                header_cells[1:], value_cells[1:], strict=False
            ):
                if col_name:
                    cat_values[col_name] = raw_val
            values[category] = cat_values

    return {"asset_name": asset_name, "sources": sources, "values": values}


def _parse_numeric(raw: str) -> float | None:
    cleaned = raw.strip().strip("*").strip()
    if cleaned in ("", "-", "—", "NC", "nc"):
        return None
    cleaned = cleaned.replace(",", ".").rstrip("%").strip()
    try:
        return round(float(cleaned), 1)
    except ValueError:
        return None


def _header_index_map(
    ws, col_start_0: int, col_end_0: int
) -> dict[str, int]:
    """{normalized header: 1-indexed openpyxl column} for a category
    range.
    """
    mapping: dict[str, int] = {}
    for col_0 in range(col_start_0, col_end_0):
        col_1 = col_0 + 1
        header = ws.cell(row=HEADER_ROW_1INDEXED, column=col_1).value
        if header is not None:
            mapping[_norm_key(str(header))] = col_1
    return mapping


def _find_reference_col(ws) -> int:
    for col_1 in range(1, ws.max_column + 1):
        header = ws.cell(row=HEADER_ROW_1INDEXED, column=col_1).value
        if header is not None and _norm_key(str(header)) == REFERENCE_HEADER:
            return col_1
    return ws.max_column


def _find_asset_row(ws, norm_target: str) -> int | None:
    for row_1 in range(DATA_START_ROW_1INDEXED, ws.max_row + 1):
        name = ws.cell(row=row_1, column=NAME_COL_0 + 1).value
        if name is not None and _norm_key(str(name)) == norm_target:
            return row_1
    return None


def resolve_base_and_target(
    allocations_dir: Path, target_date: str
) -> tuple[Path | None, Path]:
    target_file = allocations_dir / f"{target_date}.xlsx"
    if target_file.exists():
        return target_file, target_file

    applicable = sorted(
        f
        for f in allocations_dir.glob("????-??-??.xlsx")
        if f.stem <= target_date
    )
    base_file = applicable[-1] if applicable else None
    return base_file, target_file


def build(
    allocations_dir: Path,
    target_date: str,
    sheets_dir: Path,
    sheet_filter: Path | None,
    dry_run: bool,
) -> dict:
    base_file, target_file = resolve_base_and_target(
        allocations_dir, target_date
    )
    if base_file is None:
        return {
            "errors": [
                f"No reference allocation workbook found in "
                f"{allocations_dir}. Create the first "
                "data/input/allocations/YYYY-MM-DD.xlsx manually (or via "
                "finance-init) before running this script."
            ]
        }

    if sheet_filter is not None:
        sheet_files = [sheet_filter]
    else:
        sheet_files = (
            sorted(sheets_dir.glob("*.md")) if sheets_dir.is_dir() else []
        )

    synthesis: dict = {
        "date": target_date,
        "target_file": str(target_file),
        "base_file": str(base_file),
        "sheets_dir": str(sheets_dir),
        "sheets_processed": len(sheet_files),
        "dry_run": dry_run,
        "assets": [],
        "errors": [],
    }

    if not sheet_files:
        synthesis["errors"].append(
            f"No markdown reports found in {sheets_dir}"
        )
        return synthesis

    # Parse every report before touching the filesystem, so a malformed
    # report can't leave behind a half-copied workbook.
    reports: list[tuple[Path, dict]] = []
    for sheet_file in sheet_files:
        try:
            text = sheet_file.read_text(encoding="utf-8")
            reports.append((sheet_file, parse_report(text)))
        except ValueError as exc:
            synthesis["errors"].append(f"{sheet_file.name}: {exc}")

    if not reports:
        synthesis["backed_up_to"] = None
        return synthesis

    backed_up = False
    if not dry_run:
        if target_file == base_file and target_file.exists():
            backup_path = target_file.with_suffix(
                target_file.suffix + ".bak"
            )
            shutil.copy2(target_file, backup_path)
            backed_up = True
        elif target_file != base_file:
            target_file.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(base_file, target_file)
    synthesis["backed_up_to"] = (
        str(target_file.with_suffix(target_file.suffix + ".bak"))
        if backed_up
        else None
    )

    wb = openpyxl.load_workbook(target_file if not dry_run else base_file)
    ws = wb.active
    assert ws is not None, "workbook has no active worksheet"

    header_maps = {
        cat: _header_index_map(ws, start, end)
        for cat, (start, end) in CATEGORY_RANGES_0INDEXED.items()
    }
    reference_col = _find_reference_col(ws)

    for sheet_file, report in reports:
        norm_target = _norm_key(report["asset_name"])
        row_1 = _find_asset_row(ws, norm_target)
        status = "updated"
        if row_1 is None:
            row_1 = ws.max_row + 1
            ws.cell(
                row=row_1, column=NAME_COL_0 + 1, value=report["asset_name"]
            )
            status = "created"

        asset_entry: dict = {
            "asset": report["asset_name"],
            "sheet": sheet_file.name,
            "status": status,
            "changes": [],
            "skipped_cells": [],
            "warnings": [],
        }

        for category, cat_values in report["values"].items():
            header_map = header_maps.get(category, {})
            for col_name, raw_val in cat_values.items():
                col_1 = header_map.get(_norm_key(col_name))
                if col_1 is None:
                    asset_entry["skipped_cells"].append(
                        {
                            "category": category,
                            "column": col_name,
                            "reason": "column not found in workbook",
                        }
                    )
                    continue
                new_val = _parse_numeric(raw_val)
                if new_val is None:
                    asset_entry["skipped_cells"].append(
                        {
                            "category": category,
                            "column": col_name,
                            "reason": f"unresolved value '{raw_val}'",
                        }
                    )
                    continue
                old_val = ws.cell(row=row_1, column=col_1).value
                old_val = (
                    float(old_val)
                    if isinstance(old_val, int | float)
                    else None
                )
                if old_val != new_val:
                    asset_entry["changes"].append(
                        {
                            "category": category,
                            "column": col_name,
                            "old": old_val,
                            "new": new_val,
                        }
                    )
                ws.cell(row=row_1, column=col_1, value=new_val)

        if report["sources"]:
            existing_ref = ws.cell(row=row_1, column=reference_col).value
            existing_urls = (
                [
                    u.strip()
                    for u in re.split(r"[,\n\r]+", str(existing_ref))
                    if u.strip()
                ]
                if existing_ref
                else []
            )
            merged = existing_urls + [
                u for u in report["sources"] if u not in existing_urls
            ]
            ws.cell(row=row_1, column=reference_col, value=", ".join(merged))
            asset_entry["sources_added"] = len(merged) - len(existing_urls)

        if status == "created":
            asset_entry["warnings"].append(
                "New row has no ISIN/yahoo_symbol/currency/ticker set — "
                "fill these in before running finance-pipeline."
            )

        synthesis["assets"].append(asset_entry)

    if not dry_run:
        wb.save(target_file)

    return synthesis


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--date",
        default=date.today().isoformat(),
        help="Target date (default: today)",
    )
    parser.add_argument(
        "--data-dir",
        default=None,
        help="Data directory (default: auto-resolved)",
    )
    parser.add_argument(
        "--allocations-dir",
        default=None,
        help="Override <data-dir>/input/allocations",
    )
    parser.add_argument(
        "--sheets-dir", default=None, help="Override <allocations-dir>/<date>"
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Process a single report instead of the whole sheets dir",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview changes without writing the workbook",
    )
    args = parser.parse_args()

    allocations_dir = (
        Path(args.allocations_dir)
        if args.allocations_dir
        else resolve_data_dir(args.data_dir) / "input" / "allocations"
    )
    sheets_dir = (
        Path(args.sheets_dir)
        if args.sheets_dir
        else allocations_dir / args.date
    )
    sheet_filter = Path(args.sheet) if args.sheet else None

    synthesis = build(
        allocations_dir, args.date, sheets_dir, sheet_filter, args.dry_run
    )
    print(json.dumps(synthesis, ensure_ascii=False, indent=2))
    if synthesis.get("errors"):
        sys.exit(1)


if __name__ == "__main__":
    main()
